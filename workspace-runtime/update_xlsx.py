import argparse
import base64
import json
import os
from openpyxl import load_workbook


def parse_args():
    parser = argparse.ArgumentParser(description="Edit an existing workbook and save a validated copy.")
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--payload-base64", required=True)
    return parser.parse_args()


def decode_payload(value):
    raw = base64.b64decode(value.encode("utf-8"))
    return json.loads(raw.decode("utf-8"))


def normalize_sheet_name(value):
    text = str(value or "").strip()
    if not text:
        raise ValueError("Worksheet name cannot be empty.")
    return text[:31]


def main():
    args = parse_args()
    input_path = os.path.abspath(args.input)
    output_path = os.path.abspath(args.output)
    payload = decode_payload(args.payload_base64)

    remove_sheets = [str(item).strip() for item in (payload.get("removeSheets") or []) if str(item).strip()]
    keep_sheets = [str(item).strip() for item in (payload.get("keepSheets") or []) if str(item).strip()]
    rename_sheets = [
        {
            "from": str(item.get("from") or "").strip(),
            "to": normalize_sheet_name(item.get("to")),
        }
        for item in (payload.get("renameSheets") or [])
        if str(item.get("from") or "").strip()
    ]

    if remove_sheets and keep_sheets:
        raise ValueError("Use either removeSheets or keepSheets, not both together.")

    keep_vba = input_path.lower().endswith(".xlsm")
    workbook = load_workbook(filename=input_path, keep_vba=keep_vba)
    original_sheet_names = list(workbook.sheetnames)

    missing_removals = [name for name in remove_sheets if name not in workbook.sheetnames]
    if missing_removals:
        raise ValueError(f"Sheets not found for removal: {', '.join(missing_removals)}")

    missing_keeps = [name for name in keep_sheets if name not in workbook.sheetnames]
    if missing_keeps:
        raise ValueError(f"Sheets not found for keepSheets: {', '.join(missing_keeps)}")

    missing_renames = [item["from"] for item in rename_sheets if item["from"] not in workbook.sheetnames]
    if missing_renames:
        raise ValueError(f"Sheets not found for rename: {', '.join(missing_renames)}")

    if keep_sheets:
        for sheet_name in list(workbook.sheetnames):
            if sheet_name not in keep_sheets:
                workbook.remove(workbook[sheet_name])
    elif remove_sheets:
        for sheet_name in remove_sheets:
            workbook.remove(workbook[sheet_name])

    if not workbook.sheetnames:
        raise ValueError("Workbook must contain at least one worksheet after editing.")

    seen_target_names = set()
    for item in rename_sheets:
        target_name = item["to"]
        if target_name in seen_target_names:
            raise ValueError(f"Duplicate rename target detected: {target_name}")
        seen_target_names.add(target_name)

    for item in rename_sheets:
        worksheet = workbook[item["from"]]
        worksheet.title = item["to"]

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    workbook.save(output_path)

    # Re-open the output workbook once to ensure the saved file is structurally valid.
    validation_workbook = load_workbook(filename=output_path, read_only=True, keep_vba=keep_vba)
    validated_sheet_names = list(validation_workbook.sheetnames)
    validation_workbook.close()

    print(json.dumps({
        "ok": True,
        "inputPath": input_path,
        "outputPath": output_path,
        "originalSheetNames": original_sheet_names,
        "sheetNames": validated_sheet_names,
        "removedSheets": remove_sheets,
        "keptSheets": keep_sheets,
        "renamedSheets": rename_sheets,
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
