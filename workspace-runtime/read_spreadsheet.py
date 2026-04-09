import argparse
import csv
import json
import os
from openpyxl import load_workbook


def parse_args():
    parser = argparse.ArgumentParser(description="Read spreadsheet content as structured JSON.")
    parser.add_argument("--input", required=True)
    parser.add_argument("--sheet-name")
    parser.add_argument("--max-rows", type=int, default=50)
    parser.add_argument("--header-row", default="true")
    parser.add_argument("--include-empty-rows", default="false")
    return parser.parse_args()


def to_bool(value):
    return str(value).strip().lower() == "true"


def normalize_cell(value):
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def row_is_empty(row):
    return all(cell in (None, "") for cell in row)


def build_sheet_result(name, rows, max_rows, header_row, include_empty_rows):
    normalized_rows = [[normalize_cell(cell) for cell in row] for row in rows]
    if not include_empty_rows:
        normalized_rows = [row for row in normalized_rows if not row_is_empty(row)]

    total_rows = len(normalized_rows)
    headers = normalized_rows[0] if header_row and total_rows > 0 else []
    data_rows = normalized_rows[1:] if header_row and total_rows > 0 else normalized_rows
    preview_rows = data_rows[:max_rows]
    column_count = max((len(row) for row in normalized_rows), default=0)

    return {
        "name": name,
        "headers": headers,
        "rows": preview_rows,
        "rowCount": len(data_rows),
        "columnCount": column_count,
        "truncated": len(data_rows) > len(preview_rows),
    }


def read_xlsx(input_path, max_rows, header_row, include_empty_rows, requested_sheet_name=None):
    workbook = load_workbook(filename=input_path, data_only=True, read_only=True)
    selected_sheet_names = [requested_sheet_name] if requested_sheet_name else workbook.sheetnames
    sheets = []

    for name in selected_sheet_names:
        worksheet = workbook[name]
        rows = list(worksheet.iter_rows(values_only=True))
        sheets.append(build_sheet_result(name, rows, max_rows, header_row, include_empty_rows))

    return {
        "type": "xlsx",
        "sheetNames": workbook.sheetnames,
        "sheets": sheets,
    }


def read_delimited(input_path, delimiter, max_rows, header_row, include_empty_rows):
    with open(input_path, "r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        rows = [row for row in reader]

    return {
        "type": "delimited",
        "sheetNames": ["Sheet1"],
        "sheets": [
            build_sheet_result("Sheet1", rows, max_rows, header_row, include_empty_rows),
        ],
    }


def main():
    args = parse_args()
    input_path = os.path.abspath(args.input)
    extension = os.path.splitext(input_path)[1].lower()
    header_row = to_bool(args.header_row)
    include_empty_rows = to_bool(args.include_empty_rows)
    max_rows = max(1, min(args.max_rows, 200))

    if extension == ".xlsx":
        result = read_xlsx(input_path, max_rows, header_row, include_empty_rows, args.sheet_name)
    elif extension == ".csv":
        result = read_delimited(input_path, ",", max_rows, header_row, include_empty_rows)
    elif extension == ".tsv":
        result = read_delimited(input_path, "\t", max_rows, header_row, include_empty_rows)
    else:
        raise ValueError(f"Unsupported spreadsheet format: {extension or '(unknown)'}")

    result["path"] = input_path
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
