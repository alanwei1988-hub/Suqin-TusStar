import argparse
import base64
import json
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def parse_args():
    parser = argparse.ArgumentParser(description="Create an xlsx workbook from JSON payload.")
    parser.add_argument("--output", required=True)
    parser.add_argument("--payload-base64", required=True)
    return parser.parse_args()


def decode_payload(value):
    raw = base64.b64decode(value.encode("utf-8"))
    return json.loads(raw.decode("utf-8"))


def normalize_sheet_name(value, index):
    text = str(value or "").strip()
    if not text:
        text = f"Sheet{index}"
    return text[:31]


def append_rows(ws, rows):
    max_lengths = {}

    for row in rows:
      ws.append(list(row))
      for column_index, cell_value in enumerate(row, start=1):
          text = "" if cell_value is None else str(cell_value)
          max_lengths[column_index] = max(max_lengths.get(column_index, 0), len(text))

    for column_index, length in max_lengths.items():
      ws.column_dimensions[get_column_letter(column_index)].width = min(max(length + 2, 10), 40)


def build_rows(sheet):
    header = sheet.get("header") or []
    rows = []

    if header:
        rows.append(header)

    records = sheet.get("records") or []
    if records:
        keys = list(header)
        for record in records:
            for key in record.keys():
                if key not in keys:
                    keys.append(key)
        if not header:
            rows.append(keys)
        rows.extend([[record.get(key) for key in keys] for record in records])

    explicit_rows = sheet.get("rows") or []
    rows.extend(explicit_rows)
    return rows


def main():
    args = parse_args()
    payload = decode_payload(args.payload_base64)
    sheets = payload.get("sheets") or []

    if not sheets:
        raise ValueError("At least one sheet definition is required.")

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for index, sheet in enumerate(sheets, start=1):
        ws = workbook.create_sheet(title=normalize_sheet_name(sheet.get("name"), index))
        rows = build_rows(sheet)

        if not rows:
            rows = [[]]

        append_rows(ws, rows)
        if len(rows) > 1:
            ws.freeze_panes = "A2"

    os.makedirs(os.path.dirname(args.output), exist_ok=True)
    workbook.save(args.output)
    print(json.dumps({
        "ok": True,
        "outputPath": args.output,
        "sheetCount": len(workbook.sheetnames),
        "sheetNames": workbook.sheetnames,
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
